import openpyxl, pathlib, re
wb = openpyxl.load_workbook('5_6_400_2700_1V-2V_fittedout1.xlsx', data_only=True)
ws = wb[wb.sheetnames[0]]
header = [c for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
amp_cols = []
for idx, h in enumerate(header[1:], start=2):
    if h is None:
        continue
    m = re.match(r'(\d+)mV_ControlWord', str(h))
    if m:
        amp_cols.append((int(m.group(1)), idx))
amp_cols.sort()
rows = []
missing = []
for r in ws.iter_rows(min_row=2, values_only=True):
    if r[0] is None:
        continue
    f = int(r[0])
    items = []
    for amp, col in amp_cols:
        v = r[col - 1]
        if v is None or str(v).strip() == '':
            missing.append((f, amp))
            code = 0
        else:
            code = int(float(v))
        items.append((amp, code))
    rows.append((f, items))
rows.sort(key=lambda x: x[0])
lines = []
for f, items in rows:
    chunk = ', '.join([f'{{{f}, {amp}, {code}}}' for amp, code in items])
    lines.append(f'  /* {f}Hz */ ' + chunk + ',')
text = '\n'.join(lines) + '\n'
pathlib.Path('full_table_from_fittedout1.txt').write_text(text, encoding='utf-8')
print('freq_count', len(rows), 'freq_minmax', rows[0][0], rows[-1][0])
print('amp_points', [a for a, _ in amp_cols])
print('missing', len(missing))
print('written', len(lines), 'lines')
